#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Извлекает СНИЛС (B1), ФИО (B2) и дату рождения (B3)
из всех файлов с именем "Персональные данные.xlsx" в указанной
директории и её вложенных папках. Результат пишет в txt:

00000000000 Фамилия Имя Отчество ДД.ММ.ГГГГ

Используется только содержимое Excel, структура папок и имена файлов,
кроме точного совпадения названия книги, не анализируются.

Запуск:
  python scripts/extract_personal_data_from_xlsx.py [ROOT_DIR] [--out OUT]

По умолчанию:
- ROOT_DIR = текущая директория
- OUT = scripts/personal_data.txt
"""

from __future__ import annotations

import os
import re
import sys
import argparse
from datetime import datetime
from typing import Iterable, List, Optional, Tuple

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None


EXPECTED_XLSX_NAME = "персональные данные.xlsx"
SNILS_DIGITS_RE = re.compile(r"\D+")


def is_target_excel(filename: str) -> bool:
    """True, если это именно "Персональные данные.xlsx" (без учёта регистра)."""
    return filename.lower() == EXPECTED_XLSX_NAME


def read_personal_data(path: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """Читает B1 (СНИЛС), B2 (ФИО) и B3 (дату рождения) из книги."""
    if openpyxl is None:
        return None, None, None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return None, None, None
        b1 = ws["B1"].value
        b2 = ws["B2"].value
        b3 = ws["B3"].value
        wb.close()
    except Exception:
        return None, None, None

    snils = None
    if b1 is not None:
        snils_raw = str(b1)
        snils_digits = SNILS_DIGITS_RE.sub("", snils_raw)
        if len(snils_digits) == 11 and snils_digits.isdigit():
            snils = snils_digits

    fio = str(b2).strip() if b2 is not None else None
    
    # Обработка даты рождения
    birth_date = None
    if b3 is not None:
        if isinstance(b3, datetime):
            birth_date = b3.strftime("%d.%m.%Y")
        else:
            # Если дата как строка, попробуем распарсить
            birth_str = str(b3).strip()
            if birth_str:
                # Попробуем разные форматы даты
                date_formats = ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"]
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.strptime(birth_str, fmt)
                        birth_date = parsed_date.strftime("%d.%m.%Y")
                        break
                    except ValueError:
                        continue
                # Если не удалось распарсить, оставляем как есть (если похоже на дату)
                if not birth_date and re.match(r'\d{1,2}\.\d{1,2}\.\d{4}', birth_str):
                    birth_date = birth_str

    return snils, fio, birth_date


def surname_from_fio(fio: Optional[str]) -> Optional[str]:
    """Извлекает фамилию из ФИО (первое слово)."""
    if not fio:
        return None
    parts = fio.strip().split()
    if not parts:
        return None
    return parts[0]


def format_personal_record(snils: str, fio: str, birth_date: str) -> str:
    """Форматирует запись в требуемом формате: СНИЛС ФИО дата_рождения."""
    return f"{snils} {fio} {birth_date}"


def find_all_personal_data_xlsx(root: str) -> Iterable[str]:
    for dirpath, _, filenames in os.walk(root):
        for fname in filenames:
            if is_target_excel(fname):
                yield os.path.join(dirpath, fname)


def write_records(records: List[str], out_path: str) -> None:
    """Записывает строки в файл."""
    with open(out_path, "w", encoding="utf-8") as f:
        for record in records:
            f.write(f"{record}\n")


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Собирает СНИЛС (B1), ФИО (B2) и дату рождения (B3) из файлов 'Персональные данные.xlsx'"
        )
    )
    parser.add_argument(
        "root",
        nargs="?",
        default=os.getcwd(),
        help="Корневая папка для сканирования (по умолчанию: текущая)",
    )
    parser.add_argument(
        "--out",
        dest="out_path",
        default=None,
        help="Путь к выходному txt (по умолчанию: scripts/personal_data.txt)",
    )

    args = parser.parse_args(argv)
    root = os.path.abspath(args.root)
    if not os.path.isdir(root):
        print(f"[ERROR] Не папка: {root}", file=sys.stderr)
        return 2

    script_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = args.out_path or os.path.join(script_dir, "personal_data.txt")

    if openpyxl is None:
        print("[ERROR] Модуль openpyxl не установлен", file=sys.stderr)
        return 3

    records: List[str] = []
    files_found = 0
    for xlsx_path in find_all_personal_data_xlsx(root):
        files_found += 1
        snils, fio, birth_date = read_personal_data(xlsx_path)
        if snils and fio and birth_date:
            record = format_personal_record(snils, fio, birth_date)
            records.append(record)

    write_records(records, out_path)

    print(f"[OK] Найдено файлов: {files_found}")
    print(f"[OK] Извлечено записей: {len(records)}")
    print(f"[OK] Результат: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

